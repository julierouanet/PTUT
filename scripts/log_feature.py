#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
log_feature.py — Enregistre une nouvelle feature dans verification/suivi_verifications.xlsx

Usage :
    python scripts/log_feature.py --nom "Nom de la feature" --desc "Description concise"
    python scripts/log_feature.py --nom "..." --desc "..." --notes "Remarques optionnelles"

Dépendance :
    pip install openpyxl

Comportement :
    - Crée le dossier verification/ s'il n'existe pas.
    - Crée le fichier Excel avec en-têtes formatés s'il n'existe pas.
    - Ajoute une ligne sans jamais écraser les données existantes.
    - Auto-incrémente l'ID (FEAT-001, FEAT-002, ...).
"""

import argparse
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERREUR] openpyxl non installe. Executer : pip install openpyxl")
    sys.exit(1)

# Forcer UTF-8 sur la sortie standard (terminal Windows cp1252)
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Configuration ──────────────────────────────────────────────────────────────

# Chemin vers la racine du projet (dossier parent de scripts/)
PROJET_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VERIFICATION_DIR = os.path.join(PROJET_ROOT, 'verification')
EXCEL_PATH = os.path.join(VERIFICATION_DIR, 'suivi_verifications.xlsx')

# Colonnes du tableau
COLONNES = [
    'ID Feature',
    'Nom de la Feature',
    'Description concise',
    "Date d'ajout",
    'Statut Vérification',
    'Vérifié par (Humain)',
    'Date de Validation',
    'Notes / Commentaires',
]

# Largeurs des colonnes (en caractères)
LARGEURS = [12, 35, 55, 14, 22, 22, 18, 45]

# Styles
STYLE_EN_TETE_FOND    = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
STYLE_EN_TETE_POLICE  = Font(color='FFFFFF', bold=True, size=11)
STYLE_EN_ATTENTE_FOND = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
STYLE_VALIDE_FOND     = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')


# ── Fonctions utilitaires ──────────────────────────────────────────────────────

def creer_classeur() -> openpyxl.Workbook:
    """Crée un nouveau classeur Excel avec les en-têtes formatés."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Suivi Features'

    # Figer la première ligne (en-têtes)
    ws.freeze_panes = 'A2'

    # Écrire les en-têtes
    for idx, (colonne, largeur) in enumerate(zip(COLONNES, LARGEURS), start=1):
        cellule = ws.cell(row=1, column=idx, value=colonne)
        cellule.font      = STYLE_EN_TETE_POLICE
        cellule.fill      = STYLE_EN_TETE_FOND
        cellule.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = largeur

    ws.row_dimensions[1].height = 32
    return wb


def prochain_id(ws) -> str:
    """Calcule le prochain ID feature (FEAT-001, FEAT-002, ...)."""
    max_num = 0
    for ligne in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        valeur = ligne[0]
        if valeur and str(valeur).upper().startswith('FEAT-'):
            try:
                num = int(str(valeur).split('-')[1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
    return f'FEAT-{max_num + 1:03d}'


def enregistrer_feature(nom: str, description: str, notes: str = '') -> str:
    """
    Ajoute une feature au fichier Excel.
    Retourne l'ID généré (ex: 'FEAT-007').
    """
    # Créer le dossier si nécessaire
    os.makedirs(VERIFICATION_DIR, exist_ok=True)

    # Ouvrir ou créer le classeur
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = creer_classeur()
        ws = wb.active

    feat_id    = prochain_id(ws)
    aujourd_hui = date.today().strftime('%Y-%m-%d')

    # Données de la nouvelle ligne
    nouvelle_ligne = [
        feat_id,                # ID Feature
        nom,                    # Nom de la Feature
        description,            # Description concise
        aujourd_hui,            # Date d'ajout
        'En attente de revue',  # Statut Vérification
        '',                     # Vérifié par (Humain)
        '',                     # Date de Validation
        notes,                  # Notes / Commentaires
    ]

    # Écrire la ligne
    prochaine_ligne = ws.max_row + 1
    for idx, valeur in enumerate(nouvelle_ligne, start=1):
        cellule = ws.cell(row=prochaine_ligne, column=idx, value=valeur)
        cellule.alignment = Alignment(vertical='top', wrap_text=True)

        # Colorer la colonne "Statut Vérification" en jaune clair
        if idx == 5:
            cellule.fill = STYLE_EN_ATTENTE_FOND

    ws.row_dimensions[prochaine_ligne].height = 28

    # Sauvegarder
    wb.save(EXCEL_PATH)
    return feat_id


def mettre_a_jour_statut(feat_id: str, statut: str, verifie_par: str = '', notes: str = '') -> bool:
    """
    Met à jour le statut d'une feature existante (usage manuel).
    Statuts valides : 'En attente de revue' | 'Validé' | 'Rejeté' | 'En cours de correction'
    Retourne True si la feature a été trouvée et mise à jour.
    """
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Fichier introuvable : {EXCEL_PATH}")
        return False

    statuts_valides = {'En attente de revue', 'Validé', 'Rejeté', 'En cours de correction'}
    if statut not in statuts_valides:
        print(f"❌ Statut invalide '{statut}'. Valeurs acceptées : {statuts_valides}")
        return False

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    for ligne in ws.iter_rows(min_row=2):
        if ligne[0].value and str(ligne[0].value).upper() == feat_id.upper():
            # Colonne 5 = Statut, 6 = Vérifié par, 7 = Date de validation, 8 = Notes
            ligne[4].value = statut
            if verifie_par:
                ligne[5].value = verifie_par
            if statut == 'Validé':
                ligne[6].value = date.today().strftime('%Y-%m-%d')
                ligne[4].fill = STYLE_VALIDE_FOND
                ligne[5].fill = STYLE_VALIDE_FOND
                ligne[6].fill = STYLE_VALIDE_FOND
            else:
                ligne[4].fill = STYLE_EN_ATTENTE_FOND
            if notes:
                ligne[7].value = notes
            wb.save(EXCEL_PATH)
            print(f"✅ {feat_id} mis à jour → statut : '{statut}'")
            return True

    print(f"❌ Feature '{feat_id}' introuvable dans le fichier.")
    return False


# ── Interface en ligne de commande ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Gestion du fichier verification/suivi_verifications.xlsx',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  # Enregistrer une nouvelle feature
  python scripts/log_feature.py --nom "Export PDF rapports" --desc "Génération PDF depuis ReportsScreen via flutter_pdfview"

  # Mettre à jour le statut (usage humain)
  python scripts/log_feature.py --update FEAT-003 --statut "Validé" --par "Lucas"
        """
    )

    # Sous-commande : ajout d'une feature (défaut)
    parser.add_argument('--nom',    help='Nom court de la feature (obligatoire pour ajout)')
    parser.add_argument('--desc',   help='Description concise, 1-2 phrases (obligatoire pour ajout)')
    parser.add_argument('--notes',  default='', help='Notes optionnelles')

    # Sous-commande : mise à jour de statut
    parser.add_argument('--update', metavar='FEAT_ID', help='ID de la feature à mettre à jour (ex: FEAT-003)')
    parser.add_argument('--statut', help="Nouveau statut : 'Validé' | 'Rejeté' | 'En cours de correction'")
    parser.add_argument('--par',    default='', help='Nom du vérificateur humain')

    args = parser.parse_args()

    # ── Mode mise à jour ──────────────────────────────────
    if args.update:
        if not args.statut:
            parser.error("--statut est requis avec --update")
        mettre_a_jour_statut(
            feat_id=args.update,
            statut=args.statut,
            verifie_par=args.par,
            notes=args.notes
        )
        return

    # ── Mode ajout ────────────────────────────────────────
    if not args.nom or not args.desc:
        parser.error("--nom et --desc sont obligatoires pour enregistrer une feature")

    feat_id = enregistrer_feature(nom=args.nom, description=args.desc, notes=args.notes)

    print(f"✅ Feature enregistrée : {feat_id} — \"{args.nom}\"")
    print(f"   Fichier : {EXCEL_PATH}")


if __name__ == '__main__':
    main()
